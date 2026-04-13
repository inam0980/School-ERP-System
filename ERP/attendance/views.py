import io
import json
import calendar
from datetime import date, timedelta

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q, Count

from accounts.decorators import role_required
from students.models import Student
from core.models import Section
from .models import Attendance, StaffAttendance
from .forms import AttendanceFilterForm, ReportFilterForm

_STAFF = ('SUPER_ADMIN', 'ADMIN', 'TEACHER', 'ACCOUNTANT', 'STAFF')
_ADMIN = ('SUPER_ADMIN', 'ADMIN', 'TEACHER')   # who can mark attendance


# ─────────────────────── TAKE ATTENDANCE (main page) ────────────────────────

@login_required
@role_required(*_ADMIN)
def take_attendance(request):
    """Renders the shell page. Student list is loaded via AJAX."""
    form = AttendanceFilterForm(request.GET or None)
    today = date.today().isoformat()
    return render(request, 'attendance/take_attendance.html', {
        'form': form,
        'today': today,
    })


# ─────────────────────── API: load students + existing records ───────────────

@login_required
@role_required(*_ADMIN)
def api_attendance_session(request):
    """
    GET /attendance/api/session/?section_id=X&date=YYYY-MM-DD
    Returns JSON with student list and any already-marked attendance for that date.
    """
    section_id = request.GET.get('section_id')
    att_date   = request.GET.get('date', date.today().isoformat())

    if not section_id:
        return JsonResponse({'error': 'section_id required'}, status=400)

    section  = get_object_or_404(Section, pk=section_id)
    students = (
        Student.objects
        .filter(section=section, is_active=True)
        .order_by('roll_number', 'full_name')
        .select_related('division', 'grade', 'section')
    )

    # Fetch existing attendance for this section on this date
    existing = {
        a.student_id: {'status': a.status, 'remarks': a.remarks}
        for a in Attendance.objects.filter(
            student__section=section,
            date=att_date,
        )
    }

    data = {
        'section': str(section),
        'date': att_date,
        'already_saved': bool(existing),
        'students': [
            {
                'id':          s.pk,
                'name':        s.full_name,
                'arabic_name': s.arabic_name,
                'student_id':  s.student_id,
                'roll':        s.roll_number or '—',
                'photo':       s.photo.url if s.photo else None,
                'status':      existing.get(s.pk, {}).get('status', 'P'),
                'remarks':     existing.get(s.pk, {}).get('remarks', ''),
            }
            for s in students
        ],
    }
    return JsonResponse(data)


# ─────────────────────── AJAX SUBMIT (bulk create/update) ───────────────────

@login_required
@role_required(*_ADMIN)
@require_http_methods(['POST'])
def submit_attendance(request):
    """
    POST /attendance/submit/
    Body JSON: { date: "YYYY-MM-DD", section_id: X, records: [{student_id, status, remarks}] }
    Uses update_or_create to handle re-submission without duplicates.
    """
    try:
        payload    = json.loads(request.body)
        att_date   = payload['date']
        records    = payload['records']   # list of {student_id, status, remarks}
    except (json.JSONDecodeError, KeyError) as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

    if not records:
        return JsonResponse({'success': False, 'error': 'No records provided'}, status=400)

    created_count = 0
    updated_count = 0
    valid_statuses = {'P', 'A', 'L', 'E'}

    for rec in records:
        student_id = rec.get('student_id')
        status     = rec.get('status', 'P')
        remarks    = rec.get('remarks', '').strip()[:255]

        if status not in valid_statuses:
            continue

        _, created = Attendance.objects.update_or_create(
            student_id=student_id,
            date=att_date,
            defaults={
                'status':    status,
                'remarks':   remarks,
                'marked_by': request.user,
            },
        )
        if created:
            created_count += 1
        else:
            updated_count += 1

    return JsonResponse({
        'success': True,
        'created': created_count,
        'updated': updated_count,
        'total':   created_count + updated_count,
        'date':    att_date,
    })


# ─────────────────────── ATTENDANCE REPORT ──────────────────────────────────

@login_required
@role_required(*_STAFF)
def attendance_report(request):
    """Filterable report showing attendance % per student."""
    form = ReportFilterForm(request.GET or None)
    rows = []
    applied = False

    if form.is_valid():
        division     = form.cleaned_data.get('division')
        grade        = form.cleaned_data.get('grade')
        section      = form.cleaned_data.get('section')
        student_name = form.cleaned_data.get('student_name', '')
        date_from    = form.cleaned_data.get('date_from')
        date_to      = form.cleaned_data.get('date_to') or date.today()

        if not date_from:
            date_from = date_to.replace(day=1)   # default: start of current month

        applied = True
        student_qs = Student.objects.filter(is_active=True).select_related(
            'division', 'grade', 'section'
        )
        if division:
            student_qs = student_qs.filter(division=division)
        if grade:
            student_qs = student_qs.filter(grade=grade)
        if section:
            student_qs = student_qs.filter(section=section)
        if student_name:
            student_qs = student_qs.filter(
                Q(full_name__icontains=student_name) |
                Q(arabic_name__icontains=student_name) |
                Q(student_id__icontains=student_name)
            )

        for student in student_qs[:200]:     # safety cap
            att_qs = Attendance.objects.filter(
                student=student,
                date__gte=date_from,
                date__lte=date_to,
            )
            total   = att_qs.count()
            present = att_qs.filter(status='P').count()
            late    = att_qs.filter(status='L').count()
            absent  = att_qs.filter(status='A').count()
            excused = att_qs.filter(status='E').count()
            pct     = round((present + late) / total * 100, 1) if total else None

            rows.append({
                'student':  student,
                'total':    total,
                'present':  present,
                'late':     late,
                'absent':   absent,
                'excused':  excused,
                'pct':      pct,
            })

        # Sort by % ascending (worst first)
        rows.sort(key=lambda r: (r['pct'] is None, r['pct'] or 0))

    return render(request, 'attendance/attendance_report.html', {
        'form':      form,
        'rows':      rows,
        'applied':   applied,
        'date_from': form.cleaned_data.get('date_from') if form.is_valid() else None,
        'date_to':   form.cleaned_data.get('date_to')   if form.is_valid() else None,
    })


# ─────────────────────── STUDENT MONTHLY CALENDAR ───────────────────────────

@login_required
@role_required(*_STAFF)
def student_calendar(request, pk):
    """Monthly calendar view showing one student's attendance."""
    student = get_object_or_404(
        Student.objects.select_related('division', 'grade', 'section'),
        pk=pk,
    )
    today = date.today()
    year  = int(request.GET.get('year',  today.year))
    month = int(request.GET.get('month', today.month))

    # Clamp
    year  = max(2020, min(year, today.year + 1))
    month = max(1,    min(month, 12))

    # Attendance records for this month
    att_qs = Attendance.objects.filter(
        student=student, date__year=year, date__month=month
    )
    att_map = {a.date.day: a for a in att_qs}

    # Build calendar weeks (list of 7 day numbers; 0 = padding)
    cal       = calendar.monthcalendar(year, month)
    month_name = calendar.month_name[month]

    # Prev / next navigation
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1

    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1

    # Summary counts for this month
    total   = att_qs.count()
    present = att_qs.filter(status='P').count()
    late    = att_qs.filter(status='L').count()
    absent  = att_qs.filter(status='A').count()
    excused = att_qs.filter(status='E').count()
    pct     = round((present + late) / total * 100, 1) if total else None

    # Pre-process weeks so the template needs no custom filters
    STATUS_LABELS = {'P': 'Present', 'A': 'Absent', 'L': 'Late', 'E': 'Excused'}
    weeks = []
    for week in cal:
        cells = []
        for day in week:
            if day == 0:
                cells.append(None)
            else:
                att = att_map.get(day)
                cells.append({
                    'day':      day,
                    'att':      att,
                    'status':   att.status if att else None,
                    'status_label': STATUS_LABELS.get(att.status, '') if att else '',
                    'remarks':  att.remarks if att else '',
                    'is_today': (day == today.day and month == today.month and year == today.year),
                })
        weeks.append(cells)

    return render(request, 'attendance/student_calendar.html', {
        'student':    student,
        'year':       year,
        'month':      month,
        'month_name': month_name,
        'weeks':      weeks,
        'prev_year':  prev_year,  'prev_month': prev_month,
        'next_year':  next_year,  'next_month': next_month,
        'today':      today,
        'stats':      {
            'total': total, 'present': present, 'late': late,
            'absent': absent, 'excused': excused, 'pct': pct,
        },
        'month_range': range(1, 13),
    })


# ─────────────────────── API: TODAY'S SCHOOL-WIDE SUMMARY ───────────────────

@login_required
def api_today_summary(request):
    """
    GET /attendance/api/summary/
    Returns JSON: total active students, how many marked, %present today.
    Used by dashboard card.
    """
    today         = date.today()
    total_active  = Student.objects.filter(is_active=True).count()
    marked_today  = Attendance.objects.filter(date=today).count()
    present_today = Attendance.objects.filter(date=today, status__in=['P', 'L']).count()
    pct = round(present_today / total_active * 100, 1) if total_active else 0

    return JsonResponse({
        'date':          today.isoformat(),
        'total_active':  total_active,
        'marked':        marked_today,
        'present':       present_today,
        'pct_present':   pct,
    })


# ─────────────────────── EXPORT TO EXCEL ────────────────────────────────────

@login_required
@role_required(*_ADMIN)
def export_attendance(request):
    """
    GET /attendance/export/?section_id=X&date_from=Y&date_to=Z
    Downloads an xlsx file.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    section_id = request.GET.get('section_id')
    date_from  = request.GET.get('date_from')
    date_to    = request.GET.get('date_to', date.today().isoformat())

    if not date_from:
        dt_to   = date.fromisoformat(date_to)
        date_from = dt_to.replace(day=1).isoformat()

    att_qs = Attendance.objects.select_related(
        'student', 'student__grade', 'student__section', 'marked_by'
    ).filter(date__gte=date_from, date__lte=date_to).order_by('date', 'student__full_name')

    if section_id:
        att_qs = att_qs.filter(student__section_id=section_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Header row styling
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ['Date', 'Student ID', 'Student Name', 'Grade', 'Section', 'Status', 'Remarks', 'Marked By']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    status_labels = {'P': 'Present', 'A': 'Absent', 'L': 'Late', 'E': 'Excused'}
    status_colors = {'P': 'C6EFCE', 'A': 'FFC7CE', 'L': 'FFEB9C', 'E': 'BDD7EE'}

    for row_idx, a in enumerate(att_qs, 2):
        data = [
            str(a.date),
            a.student.student_id,
            a.student.full_name,
            str(a.student.grade),
            str(a.student.section),
            status_labels.get(a.status, a.status),
            a.remarks,
            a.marked_by.full_name if a.marked_by else '—',
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col == 6:   # Status column — colour code
                cell.fill = PatternFill("solid", fgColor=status_colors.get(a.status, 'FFFFFF'))

    # Auto-size columns
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    fname = f"attendance_{date_from}_to_{date_to}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response
