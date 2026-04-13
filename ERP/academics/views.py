import io
import csv
import json
from collections import defaultdict
from datetime import date

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Avg, Count
from django.views.decorators.http import require_POST
from django.utils import timezone

from accounts.decorators import role_required
from students.models import Student
from core.models import Subject, Section, AcademicYear
from .models import ExamType, Exam, Mark, GradeConfig, ReportCard, TERM_CHOICES
from .forms import (
    ExamTypeForm, ExamForm, GradeConfigForm, ReportCardFilterForm,
)

_ADMIN   = ('SUPER_ADMIN', 'ADMIN')
_TEACHER = ('SUPER_ADMIN', 'ADMIN', 'TEACHER')
_STAFF   = ('SUPER_ADMIN', 'ADMIN', 'TEACHER', 'ACCOUNTANT', 'STAFF')


# ════════════════════ EXAM TYPE CRUD ════════════════════

@login_required
@role_required(*_ADMIN)
def exam_type_list(request):
    types = ExamType.objects.all()
    return render(request, 'academics/exam_type_list.html', {'types': types})


@login_required
@role_required(*_ADMIN)
def exam_type_form(request, pk=None):
    instance = get_object_or_404(ExamType, pk=pk) if pk else None
    form     = ExamTypeForm(request.POST or None, instance=instance)
    if form.is_valid():
        form.save()
        messages.success(request, "Exam type saved.")
        return redirect('academics:exam_type_list')
    return render(request, 'academics/exam_type_form.html', {
        'form':  form,
        'title': 'Edit Exam Type' if instance else 'Add Exam Type',
    })


@login_required
@role_required(*_ADMIN)
@require_POST
def exam_type_delete(request, pk):
    et = get_object_or_404(ExamType, pk=pk)
    et.delete()
    messages.success(request, "Exam type deleted.")
    return redirect('academics:exam_type_list')


# ════════════════════ EXAM CRUD ════════════════════

@login_required
@role_required(*_TEACHER)
def exam_list(request):
    qs = Exam.objects.select_related(
        'exam_type', 'subject', 'section', 'academic_year'
    )
    section_id = request.GET.get('section')
    subject_id = request.GET.get('subject')
    term       = request.GET.get('term')
    year_id    = request.GET.get('year')

    if section_id:
        qs = qs.filter(section_id=section_id)
    if subject_id:
        qs = qs.filter(subject_id=subject_id)
    if term:
        qs = qs.filter(term=term)
    if year_id:
        qs = qs.filter(academic_year_id=year_id)

    return render(request, 'academics/exam_list.html', {
        'exams':    qs.order_by('-date')[:200],
        'sections': Section.objects.select_related('grade').all(),
        'subjects': Subject.objects.filter(is_active=True),
        'years':    AcademicYear.objects.all(),
        'terms':    TERM_CHOICES,
        'filters':  request.GET,
    })


@login_required
@role_required(*_ADMIN)
def exam_form(request, pk=None):
    instance = get_object_or_404(Exam, pk=pk) if pk else None
    form     = ExamForm(request.POST or None, instance=instance)
    if form.is_valid():
        obj = form.save(commit=False)
        if not instance:
            obj.created_by = request.user
        obj.save()
        messages.success(request, "Exam saved.")
        return redirect('academics:exam_list')
    return render(request, 'academics/exam_form.html', {
        'form':  form,
        'title': 'Edit Exam' if instance else 'Add Exam',
        'exam':  instance,
    })


@login_required
@role_required(*_ADMIN)
@require_POST
def exam_delete(request, pk):
    exam = get_object_or_404(Exam, pk=pk)
    exam.delete()
    messages.success(request, "Exam deleted.")
    return redirect('academics:exam_list')


# ════════════════════ MARKS ENTRY ════════════════════

@login_required
@role_required(*_TEACHER)
def marks_entry(request, exam_pk):
    exam = get_object_or_404(
        Exam.objects.select_related('subject', 'section', 'exam_type', 'academic_year'),
        pk=exam_pk,
    )
    students = Student.objects.filter(
        section=exam.section, is_active=True
    ).order_by('roll_number', 'full_name')

    existing = {
        m.student_id: m
        for m in Mark.objects.filter(exam=exam).select_related('entered_by', 'approved_by')
    }

    has_approved = any(m.status == 'approved' for m in existing.values())
    is_admin     = request.user.role in ('SUPER_ADMIN', 'ADMIN')
    locked       = has_approved and not is_admin

    if request.method == 'POST' and not locked:
        action     = request.POST.get('action', 'draft')
        new_status = 'submitted' if action == 'submit' else 'draft'

        for student in students:
            absent  = request.POST.get(f'absent_{student.pk}') == 'on'
            raw     = request.POST.get(f'marks_{student.pk}', '').strip()
            remarks = request.POST.get(f'remarks_{student.pk}', '').strip()[:200]

            obtained = None
            if not absent and raw:
                try:
                    obtained = float(raw)
                    obtained = max(0, min(float(exam.total_marks), obtained))
                except ValueError:
                    pass

            mark_obj = existing.get(student.pk)
            if mark_obj:
                if mark_obj.status == 'approved' and not is_admin:
                    continue
                mark_obj.obtained_marks = obtained
                mark_obj.is_absent      = absent
                mark_obj.remarks        = remarks
                mark_obj.entered_by     = request.user
                if mark_obj.status != 'approved':
                    mark_obj.status = new_status
                mark_obj.save()
            else:
                Mark.objects.create(
                    student     = student,
                    exam        = exam,
                    obtained_marks = obtained,
                    is_absent   = absent,
                    remarks     = remarks,
                    status      = new_status,
                    entered_by  = request.user,
                )

        label = 'submitted for approval' if action == 'submit' else 'saved as draft'
        messages.success(request, f"Marks {label}.")
        return redirect('academics:marks_entry', exam_pk=exam.pk)

    rows = []
    for s in students:
        m = existing.get(s.pk)
        rows.append({
            'student':     s,
            'mark':        m,
            'obtained':    m.obtained_marks if m else '',
            'is_absent':   m.is_absent if m else False,
            'remarks':     m.remarks if m else '',
            'status':      m.status if m else 'draft',
            'pct':         m.get_percentage() if m else None,
            'letter':      m.get_letter_grade() if m else '—',
            'passed':      m.is_passed() if m else None,
            'is_approved': (m.status == 'approved') if m else False,
        })

    return render(request, 'academics/marks_entry.html', {
        'exam':         exam,
        'rows':         rows,
        'locked':       locked,
        'has_approved': has_approved,
        'is_admin':     is_admin,
        'total_marks':  exam.total_marks,
    })


# ════════════════════ APPROVAL ════════════════════

@login_required
@role_required(*_ADMIN)
def marks_approval(request):
    pending = (
        Exam.objects
        .filter(marks__status='submitted')
        .distinct()
        .select_related('subject', 'section', 'exam_type', 'academic_year')
        .annotate(pending_count=Count('marks', filter=Q(marks__status='submitted')))
    )
    return render(request, 'academics/marks_approval.html', {'pending_exams': pending})


@login_required
@role_required(*_ADMIN)
@require_POST
def approve_marks(request, exam_pk):
    exam    = get_object_or_404(Exam, pk=exam_pk)
    updated = Mark.objects.filter(exam=exam, status='submitted').update(
        status      = 'approved',
        approved_by = request.user,
        approved_at = timezone.now(),
    )
    messages.success(request, f"Approved {updated} mark(s) for {exam}.")
    return redirect('academics:marks_approval')


@login_required
@role_required(*_ADMIN)
@require_POST
def unlock_marks(request, exam_pk):
    exam    = get_object_or_404(Exam, pk=exam_pk)
    updated = Mark.objects.filter(exam=exam, status='approved').update(status='submitted')
    messages.success(request, f"Unlocked {updated} mark(s) — teacher can now edit.")
    return redirect('academics:marks_entry', exam_pk=exam.pk)


# ════════════════════ EXAM RESULTS (read-only) ════════════════════

@login_required
@role_required(*_STAFF)
def exam_results(request, exam_pk):
    exam = get_object_or_404(
        Exam.objects.select_related('subject', 'section', 'exam_type', 'academic_year'),
        pk=exam_pk,
    )
    marks = (
        Mark.objects.filter(exam=exam)
        .select_related('student')
        .order_by('student__roll_number', 'student__full_name')
    )
    rows = []
    for m in marks:
        rows.append({
            'student': m.student,
            'mark':    m,
            'pct':     m.get_percentage(),
            'letter':  m.get_letter_grade(),
            'passed':  m.is_passed(),
        })

    total      = len(rows)
    absent_cnt = sum(1 for r in rows if r['mark'].is_absent)
    passed_cnt = sum(1 for r in rows if r['passed'])
    scored_pcts = [r['pct'] for r in rows if r['pct'] is not None]
    avg_pct    = round(sum(scored_pcts) / len(scored_pcts), 1) if scored_pcts else None

    return render(request, 'academics/exam_results.html', {
        'exam':    exam,
        'rows':    rows,
        'total':   total,
        'passed':  passed_cnt,
        'failed':  total - passed_cnt - absent_cnt,
        'absent':  absent_cnt,
        'avg_pct': avg_pct,
    })


# ════════════════════ REPORT CARD ════════════════════

def _compute_report_data(student, academic_year, term):
    """Return computed per-subject averages and overall GPA for report card."""
    marks = Mark.objects.filter(
        student=student,
        exam__academic_year=academic_year,
        exam__term=term,
        status='approved',
    ).select_related('exam', 'exam__subject', 'exam__exam_type')

    by_subject = defaultdict(list)
    for m in marks:
        by_subject[m.exam.subject].append(m)

    def _letter(p):
        if p is None:  return 'AB'
        if p >= 90:    return 'A+'
        if p >= 85:    return 'A'
        if p >= 80:    return 'B+'
        if p >= 75:    return 'B'
        if p >= 70:    return 'C+'
        if p >= 65:    return 'C'
        if p >= 60:    return 'D'
        return 'F'

    GPA_MAP = {'A+': 4.0, 'A': 4.0, 'B+': 3.5, 'B': 3.0,
               'C+': 2.5, 'C': 2.0, 'D': 1.0, 'F': 0.0, 'AB': 0.0}

    results = []
    gpa_vals = []
    for subj, subj_marks in by_subject.items():
        wsum = wtotal = 0
        absent = False
        for m in subj_marks:
            w   = m.exam.exam_type.weight_percentage
            pct = m.get_percentage()
            if pct is not None:
                wsum   += pct * w
                wtotal += w
            if m.is_absent:
                absent = True
        avg  = round(wsum / wtotal, 1) if wtotal else None
        let  = _letter(avg)
        gpa  = GPA_MAP[let]
        gpa_vals.append(gpa)
        results.append({
            'subject':  subj,
            'avg_pct':  avg,
            'letter':   let,
            'gpa':      gpa,
            'passed':   (avg is not None and avg >= 60),
            'absent':   absent,
        })

    results.sort(key=lambda r: r['subject'].name)
    overall_gpa    = round(sum(gpa_vals) / len(gpa_vals), 2) if gpa_vals else None
    overall_passed = all(r['passed'] for r in results) if results else False
    return {
        'subjects':     results,
        'overall_gpa':  overall_gpa,
        'passed':       overall_passed,
    }


@login_required
@role_required(*_STAFF)
def report_card_view(request, student_pk):
    student = get_object_or_404(
        Student.objects.select_related('division', 'grade', 'section', 'academic_year'),
        pk=student_pk,
    )
    form = ReportCardFilterForm(request.GET or None)
    data = None
    if form.is_valid():
        data = _compute_report_data(
            student,
            form.cleaned_data['academic_year'],
            form.cleaned_data['term'],
        )
    return render(request, 'academics/report_card.html', {
        'student': student,
        'form':    form,
        'data':    data,
    })


@login_required
@role_required(*_STAFF)
def report_card_pdf(request, student_pk):
    student = get_object_or_404(
        Student.objects.select_related('division', 'grade', 'section', 'academic_year'),
        pk=student_pk,
    )
    year_id = request.GET.get('academic_year')
    term    = request.GET.get('term')
    if not year_id or not term:
        messages.error(request, "Select academic year and term first.")
        return redirect('academics:report_card', student_pk=student_pk)

    year = get_object_or_404(AcademicYear, pk=year_id)
    data = _compute_report_data(student, year, term)
    context = {'student': student, 'data': data, 'year': year, 'term': term}

    html_content = render(request, 'academics/report_card_print.html', context).content.decode()
    try:
        from xhtml2pdf import pisa
        buf = io.BytesIO()
        pisa.CreatePDF(html_content, dest=buf)
        buf.seek(0)
        fname    = f"report_{student.student_id}_{year.name}_{term}.pdf"
        response = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{fname}"'
        return response
    except ImportError:
        messages.warning(request, "PDF library not installed — showing print-friendly HTML.")
        return render(request, 'academics/report_card_print.html', context)


@login_required
@role_required(*_ADMIN)
def bulk_report_cards(request):
    form  = ReportCardFilterForm(request.GET or None)
    cards = []
    if form.is_valid():
        section  = form.cleaned_data['section']
        year     = form.cleaned_data['academic_year']
        term     = form.cleaned_data['term']
        students = Student.objects.filter(
            section=section, is_active=True
        ).select_related('division', 'grade', 'section', 'academic_year')
        for s in students:
            d = _compute_report_data(s, year, term)
            if d['subjects']:
                cards.append({'student': s, 'data': d})
    return render(request, 'academics/bulk_report_cards.html', {
        'form': form, 'cards': cards,
    })


# ════════════════════ NOOR CSV EXPORT ════════════════════

@login_required
@role_required(*_ADMIN)
def noor_export(request):
    section_id = request.GET.get('section')
    year_id    = request.GET.get('year')
    term       = request.GET.get('term')

    if not all([section_id, year_id, term]):
        messages.error(request, "Please select section, year, and term for Noor export.")
        return redirect('academics:exam_list')

    section  = get_object_or_404(Section, pk=section_id)
    year     = get_object_or_404(AcademicYear, pk=year_id)
    students = Student.objects.filter(section=section, is_active=True).order_by('roll_number', 'full_name')
    subjects = Subject.objects.filter(
        grade=section.grade, division=section.grade.division, is_active=True
    ).order_by('name')

    marks_qs = Mark.objects.filter(
        student__section=section,
        exam__academic_year=year,
        exam__term=term,
        status='approved',
    ).select_related('exam', 'exam__subject', 'exam__exam_type')

    # Weighted average per (student, subject)
    scored_map = defaultdict(lambda: defaultdict(lambda: {'wsum': 0.0, 'wtotal': 0}))
    for m in marks_qs:
        pct = m.get_percentage()
        if pct is not None:
            w = m.exam.exam_type.weight_percentage
            scored_map[m.student_id][m.exam.subject_id]['wsum']   += pct * w
            scored_map[m.student_id][m.exam.subject_id]['wtotal'] += w

    def _letter(p):
        if p is None: return 'AB'
        if p >= 90:   return 'A+'
        if p >= 85:   return 'A'
        if p >= 80:   return 'B+'
        if p >= 75:   return 'B'
        if p >= 70:   return 'C+'
        if p >= 65:   return 'C'
        if p >= 60:   return 'D'
        return 'F'

    GPA_MAP = {'A+': 4.0, 'A': 4.0, 'B+': 3.5, 'B': 3.0,
               'C+': 2.5, 'C': 2.0, 'D': 1.0, 'F': 0.0, 'AB': 0.0}

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    fname    = f"noor_{section}_{year.name}_{term}.csv"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    response.write('\ufeff')   # BOM for Arabic in Excel

    writer = csv.writer(response)
    header = ['Student ID', 'Full Name (EN)', 'Full Name (AR)', 'Roll No']
    for subj in subjects:
        header += [subj.name, f'{subj.name} Grade']
    header += ['Overall GPA', 'Result']
    writer.writerow(header)

    for student in students:
        row = [student.student_id, student.full_name, student.arabic_name or '', student.roll_number or '']
        gpas       = []
        all_passed = True
        for subj in subjects:
            d   = scored_map[student.pk][subj.pk]
            avg = round(d['wsum'] / d['wtotal'], 1) if d['wtotal'] else None
            let = _letter(avg)
            row += [avg if avg is not None else 'AB', let]
            gpas.append(GPA_MAP[let])
            if let in ('F', 'AB'):
                all_passed = False
        overall = round(sum(gpas) / len(gpas), 2) if gpas else ''
        row += [overall, 'Pass' if all_passed else 'Fail']
        writer.writerow(row)

    return response


# ════════════════════ EXCEL MARKS EXPORT ════════════════════

@login_required
@role_required(*_STAFF)
def export_marks_excel(request, exam_pk):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    exam  = get_object_or_404(Exam, pk=exam_pk)
    marks = Mark.objects.filter(exam=exam).select_related('student').order_by(
        'student__roll_number', 'student__full_name'
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Marks'

    head_fill = PatternFill('solid', fgColor='1e3a5f')
    head_font = Font(color='FFFFFF', bold=True)
    headers   = ['Roll No', 'Student ID', 'Name', 'Arabic Name',
                 f'Marks / {exam.total_marks}', '% Score', 'Grade', 'Status', 'Absent', 'Remarks']

    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.fill      = head_fill
        cell.font      = head_font
        cell.alignment = Alignment(horizontal='center')

    pass_fill = PatternFill('solid', fgColor='d1fae5')
    fail_fill = PatternFill('solid', fgColor='fee2e2')
    ab_fill   = PatternFill('solid', fgColor='fef9c3')

    for row_idx, m in enumerate(marks, 2):
        pct    = m.get_percentage()
        letter = m.get_letter_grade()
        passed = m.is_passed()
        if m.is_absent:
            fill = ab_fill
        elif passed:
            fill = pass_fill
        else:
            fill = fail_fill

        row_data = [
            m.student.roll_number or '',
            m.student.student_id,
            m.student.full_name,
            m.student.arabic_name or '',
            float(m.obtained_marks) if m.obtained_marks is not None else 'Absent',
            f'{pct}%' if pct is not None else 'AB',
            letter,
            m.get_status_display(),
            'Yes' if m.is_absent else 'No',
            m.remarks or '',
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell      = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill

    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname    = f"marks_{exam.pk}_{exam.subject.name}_{exam.term}.xlsx"
    response = HttpResponse(buf.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


# ════════════════════ GRADE CONFIG CRUD ════════════════════

@login_required
@role_required(*_ADMIN)
def grade_config_list(request):
    configs = GradeConfig.objects.select_related('grade', 'grade__division').all()
    return render(request, 'academics/grade_config_list.html', {'configs': configs})


@login_required
@role_required(*_ADMIN)
def grade_config_form(request, pk=None):
    instance = get_object_or_404(GradeConfig, pk=pk) if pk else None
    form     = GradeConfigForm(request.POST or None, instance=instance)
    if form.is_valid():
        form.save()
        messages.success(request, "Grade configuration saved.")
        return redirect('academics:grade_config_list')
    return render(request, 'academics/grade_config_form.html', {
        'form':  form,
        'title': 'Edit Grade Config' if instance else 'Add Grade Config',
    })


# ════════════════════ JSON APIS ════════════════════

@login_required
def api_subjects_by_section(request):
    section_id = request.GET.get('section_id')
    if not section_id:
        return JsonResponse({'subjects': []})
    try:
        section  = Section.objects.select_related('grade', 'grade__division').get(pk=section_id)
        subjects = list(Subject.objects.filter(
            grade=section.grade,
            division=section.grade.division,
            is_active=True,
        ).values('id', 'name', 'code'))
        return JsonResponse({'subjects': subjects})
    except Section.DoesNotExist:
        return JsonResponse({'subjects': []})


@login_required
def api_exams_by_subject_section(request):
    section_id = request.GET.get('section_id')
    subject_id = request.GET.get('subject_id')
    if not (section_id and subject_id):
        return JsonResponse({'exams': []})
    exams = list(Exam.objects.filter(
        section_id=section_id, subject_id=subject_id
    ).values('id', 'name', 'total_marks', 'term'))
    return JsonResponse({'exams': exams})
