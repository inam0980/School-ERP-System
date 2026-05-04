import csv
import io
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum, Count
from django.views.decorators.http import require_POST

from accounts.decorators import role_required
from .models import Student, StudentDocument, Sibling, AuthorizedPickup
from .forms import StudentForm, DocumentUploadForm, StudentFilterForm, SiblingForm, AuthorizedPickupForm
from fees.models import ExternalCandidate, ExternalCandidatePayment
from core.models import Grade, Division, Board

_ADMIN   = ('SUPER_ADMIN', 'ADMIN')
_STAFF   = ('SUPER_ADMIN', 'ADMIN', 'TEACHER', 'ACCOUNTANT', 'STAFF')


# ────────────────────────── STUDENT HUB ──────────────────────────

@login_required
@role_required(*_STAFF)
def student_hub(request):
    """Landing page: choose Regular Students or External Candidates."""
    regular_count  = Student.objects.count()
    external_count = ExternalCandidate.objects.count()
    return render(request, 'students/student_hub.html', {
        'regular_count':  regular_count,
        'external_count': external_count,
    })


# ────────────────────────── REGULAR STUDENT LIST ──────────────────────────

@login_required
@role_required(*_STAFF)
def student_list(request):
    form = StudentFilterForm(request.GET or None)
    qs   = Student.objects.select_related('division', 'grade', 'section', 'academic_year').all()

    if form.is_valid():
        q         = form.cleaned_data.get('q', '')
        division  = form.cleaned_data.get('division')
        grade     = form.cleaned_data.get('grade')
        section   = form.cleaned_data.get('section')
        gender    = form.cleaned_data.get('gender')
        is_active = form.cleaned_data.get('is_active')

        if q:
            qs = qs.filter(
                Q(full_name__icontains=q) |
                Q(arabic_name__icontains=q) |
                Q(student_id__icontains=q) |
                Q(guardian_phone__icontains=q)
            )
        if division:
            qs = qs.filter(division=division)
        if grade:
            qs = qs.filter(grade=grade)
        if section:
            qs = qs.filter(section=section)
        if gender:
            qs = qs.filter(gender=gender)
        if is_active == '1':
            qs = qs.filter(is_active=True)
        elif is_active == '0':
            qs = qs.filter(is_active=False)

        citizenship = form.cleaned_data.get('citizenship')
        if citizenship == 'saudi':
            qs = qs.filter(nationality='Saudi')
        elif citizenship == 'expat':
            qs = qs.exclude(nationality='Saudi')

    total       = qs.count()
    saudi_count = qs.filter(nationality='Saudi').count()
    expat_count = total - saudi_count
    return render(request, 'students/student_list.html', {
        'students':     qs,
        'form':         form,
        'total':        total,
        'saudi_count':  saudi_count,
        'expat_count':  expat_count,
    })


# ────────────────────────── EXPORT CSV ──────────────────────────

@login_required
@role_required(*_STAFF)
def student_export_csv(request):
    """Export the currently filtered student list as a CSV file."""
    form = StudentFilterForm(request.GET or None)
    qs   = Student.objects.select_related('division', 'grade', 'section', 'academic_year').all()

    if form.is_valid():
        q         = form.cleaned_data.get('q', '')
        division  = form.cleaned_data.get('division')
        grade     = form.cleaned_data.get('grade')
        section   = form.cleaned_data.get('section')
        gender    = form.cleaned_data.get('gender')
        is_active = form.cleaned_data.get('is_active')

        if q:
            qs = qs.filter(
                Q(full_name__icontains=q) |
                Q(arabic_name__icontains=q) |
                Q(student_id__icontains=q) |
                Q(guardian_phone__icontains=q)
            )
        if division:
            qs = qs.filter(division=division)
        if grade:
            qs = qs.filter(grade=grade)
        if section:
            qs = qs.filter(section=section)
        if gender:
            qs = qs.filter(gender=gender)
        if is_active == '1':
            qs = qs.filter(is_active=True)
        elif is_active == '0':
            qs = qs.filter(is_active=False)

        citizenship = form.cleaned_data.get('citizenship')
        if citizenship == 'saudi':
            qs = qs.filter(nationality='Saudi')
        elif citizenship == 'expat':
            qs = qs.exclude(nationality='Saudi')

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="students.csv"'
    response.write('\ufeff')  # BOM for Excel UTF-8 compatibility

    writer = csv.writer(response)
    writer.writerow([
        'Student ID', 'Full Name', 'Arabic Name', 'Gender', 'Date of Birth',
        'Nationality', 'ID Type', 'National ID', 'Division', 'Grade', 'Section',
        'Academic Year', 'Roll No.', 'Enrollment Type', 'Admission Date', 'Active',
        'Father Name', 'Mother Name', 'Guardian Phone', 'Guardian Email',
        'Address',
    ])
    for s in qs:
        writer.writerow([
            s.student_id, s.full_name, s.arabic_name,
            s.get_gender_display(), s.dob,
            s.nationality, s.get_id_type_display(), s.national_id,
            s.division, s.grade, s.section, s.academic_year,
            s.roll_number, s.get_enrollment_type_display(), s.admission_date,
            'Yes' if s.is_active else 'No',
            s.father_name, s.mother_name, s.guardian_phone, s.guardian_email,
            s.address,
        ])
    return response


# ────────────────────────── ADD ──────────────────────────

def _save_siblings_from_post(post, student):
    """Parse sibling array fields from POST and bulk-create Sibling records."""
    names    = post.getlist('sibling_full_name[]')
    relations = post.getlist('sibling_relation[]')
    dobs     = post.getlist('sibling_dob[]')
    schools  = post.getlist('sibling_school[]')
    levels   = post.getlist('sibling_level[]')
    to_create = []
    for i, name in enumerate(names):
        name = name.strip()
        if not name:
            continue
        relation = relations[i] if i < len(relations) else ''
        if relation not in ('BROTHER', 'SISTER'):
            continue
        dob = dobs[i].strip() if i < len(dobs) else ''
        to_create.append(Sibling(
            student=student,
            full_name=name,
            relation=relation,
            dob=dob or None,
            current_school=(schools[i].strip() if i < len(schools) else ''),
            educational_level=(levels[i].strip() if i < len(levels) else ''),
        ))
    if to_create:
        Sibling.objects.bulk_create(to_create)


def _save_pickups_from_post(post, student):
    """Parse pickup array fields from POST and bulk-create AuthorizedPickup records."""
    names     = post.getlist('pickup_full_name[]')
    relations = post.getlist('pickup_relation[]')
    phones    = post.getlist('pickup_phone[]')
    to_create = []
    for i, name in enumerate(names):
        name = name.strip()
        if not name:
            continue
        to_create.append(AuthorizedPickup(
            student=student,
            full_name=name,
            relation=(relations[i].strip() if i < len(relations) else ''),
            phone=(phones[i].strip() if i < len(phones) else ''),
        ))
    if to_create:
        AuthorizedPickup.objects.bulk_create(to_create)


@login_required
@role_required(*_ADMIN)
def student_add(request):
    candidate_id = request.GET.get('candidate_id')
    candidate = None
    initial = {}

    if candidate_id:
        from fees.models import ExternalCandidate
        candidate = get_object_or_404(ExternalCandidate, pk=candidate_id)
        if candidate.status == ExternalCandidate.STATUS_APPROVED and candidate.enrolled_student:
            messages.warning(request, "This candidate is already approved and enrolled.")
            return redirect('students:detail', pk=candidate.enrolled_student.pk)

        initial = {
            'full_name': candidate.full_name,
            'arabic_name': candidate.arabic_name,
            'guardian_phone': candidate.phone,
            'nationality': candidate.nationality,
            'national_id': candidate.id_number,
            'grade': candidate.grade_applying,
            'enrollment_type': Student.NEW,
        }

    form = StudentForm(request.POST or None, request.FILES or None, initial=initial)
    if form.is_valid():
        student = form.save(commit=False)
        student.created_by = request.user
        student.save()
        _save_siblings_from_post(request.POST, student)
        _save_pickups_from_post(request.POST, student)

        if candidate:
            candidate.status = ExternalCandidate.STATUS_APPROVED
            candidate.enrolled_student = student
            candidate.save()

        messages.success(request, f"Student {student.full_name} added (ID: {student.student_id}). Please upload identity documents (National ID / Iqama / Passport) below.")
        return redirect('students:detail', pk=student.pk)
    return render(request, 'students/student_form.html', {'form': form, 'title': 'Add Student / إضافة طالب'})


# ────────────────────────── DETAIL ──────────────────────────

@login_required
@role_required(*_STAFF)
def student_detail(request, pk):
    student = get_object_or_404(Student.objects.select_related(
        'division', 'grade', 'section', 'academic_year', 'created_by'
    ), pk=pk)
    doc_form     = DocumentUploadForm()
    sibling_form = SiblingForm()
    pickup_form  = AuthorizedPickupForm()
    return render(request, 'students/student_detail.html', {
        'student':      student,
        'doc_form':     doc_form,
        'documents':    student.documents.all(),
        'sibling_form': sibling_form,
        'siblings':     student.siblings.all(),
        'pickup_form':  pickup_form,
        'pickups':      student.authorized_pickups.all(),
    })


# ────────────────────────── EDIT ──────────────────────────

@login_required
@role_required(*_ADMIN)
def student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk)
    form    = StudentForm(request.POST or None, request.FILES or None, instance=student)
    if form.is_valid():
        form.save()
        # Replace all siblings with what was submitted
        student.siblings.all().delete()
        _save_siblings_from_post(request.POST, student)
        # Replace all pickups with what was submitted
        student.authorized_pickups.all().delete()
        _save_pickups_from_post(request.POST, student)
        messages.success(request, "Student updated successfully.")
        return redirect('students:detail', pk=student.pk)
    return render(request, 'students/student_form.html', {
        'form': form,
        'student': student,
        'existing_siblings': student.siblings.all(),
        'existing_pickups':  student.authorized_pickups.all(),
        'title': f'Edit: {student.full_name}',
    })


# ────────────────────────── SOFT DELETE ──────────────────────────

@login_required
@role_required(*_ADMIN)
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        student.is_active = False
        student.save(update_fields=['is_active'])
        messages.success(request, f"Student {student.full_name} has been deactivated.")
        return redirect('students:list')
    return render(request, 'students/student_confirm_delete.html', {'student': student})


# ────────────────────────── DOCUMENT UPLOAD ──────────────────────────

@login_required
@role_required(*_ADMIN)
def document_upload(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = DocumentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.student     = student
            doc.uploaded_by = request.user
            doc.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'id':       doc.pk,
                    'doc_type': doc.get_doc_type_display(),
                    'filename': doc.filename,
                    'url':      doc.file.url,
                })
            messages.success(request, "Document uploaded.")
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
            messages.error(request, "Upload failed. Please check the form.")
    return redirect('students:detail', pk=pk)


@login_required
@role_required(*_ADMIN)
def document_delete(request, doc_pk):
    doc = get_object_or_404(StudentDocument, pk=doc_pk)
    student_pk = doc.student.pk
    if request.method == 'POST':
        doc.file.delete(save=False)
        doc.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': True})
        messages.success(request, "Document deleted.")
    return redirect('students:detail', pk=student_pk)


# ────────────────────────── SIBLING ADD / DELETE ──────────────────────────

@login_required
@role_required(*_ADMIN)
@require_POST
def sibling_add(request, pk):
    student = get_object_or_404(Student, pk=pk)
    form = SiblingForm(request.POST)
    if form.is_valid():
        sibling = form.save(commit=False)
        sibling.student = student
        sibling.save()
        messages.success(request, "Sibling added.")
    else:
        messages.error(request, "Please correct the sibling form errors.")
    return redirect('students:detail', pk=pk)


@login_required
@role_required(*_ADMIN)
@require_POST
def sibling_delete(request, sibling_pk):
    sibling = get_object_or_404(Sibling, pk=sibling_pk)
    student_pk = sibling.student.pk
    sibling.delete()
    messages.success(request, "Sibling removed.")
    return redirect('students:detail', pk=student_pk)


# ────────────────────────── AUTHORIZED PICKUP ADD / DELETE ───────────

@login_required
@role_required(*_ADMIN)
@require_POST
def pickup_add(request, pk):
    student = get_object_or_404(Student, pk=pk)
    form = AuthorizedPickupForm(request.POST)
    if form.is_valid():
        pickup = form.save(commit=False)
        pickup.student = student
        pickup.save()
        messages.success(request, "Authorized person added.")
    else:
        messages.error(request, "Please correct the form errors.")
    return redirect('students:detail', pk=pk)


@login_required
@role_required(*_ADMIN)
@require_POST
def pickup_delete(request, pickup_pk):
    pickup = get_object_or_404(AuthorizedPickup, pk=pickup_pk)
    student_pk = pickup.student.pk
    pickup.delete()
    messages.success(request, "Authorized person removed.")
    return redirect('students:detail', pk=student_pk)


# ────────────────────────── ID CARD PRINT ──────────────────────────

@login_required
def student_id_card(request, pk):
    student = get_object_or_404(Student, pk=pk)
    return render(request, 'students/id_card.html', {'student': student})


# ────────────────────────── EXCEL IMPORT ──────────────────────────

@login_required
@role_required(*_ADMIN)
def student_import(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        try:
            import openpyxl
            from core.models import AcademicYear, Division, Grade, Section
            from datetime import date

            wb   = openpyxl.load_workbook(request.FILES['excel_file'])
            ws   = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))

            created = 0
            errors  = []

            for i, row in enumerate(rows, start=2):
                if not row or not row[0]:
                    continue
                try:
                    # Expected columns (1-indexed):
                    # 1:full_name 2:arabic_name 3:dob(YYYY-MM-DD) 4:gender(M/F)
                    # 5:nationality 6:division_name 7:grade_name 8:section_name
                    # 9:academic_year_name 10:father_name 11:mother_name
                    # 12:guardian_phone 13:guardian_email 14:admission_date
                    full_name    = str(row[0]).strip()
                    arabic_name  = str(row[1]).strip() if row[1] else ''
                    dob          = row[2] if isinstance(row[2], date) else date.fromisoformat(str(row[2]))
                    gender       = str(row[3]).strip().upper()
                    nationality  = str(row[4]).strip() if row[4] else 'Saudi'
                    div_name     = str(row[5]).strip().upper()
                    grade_name   = str(row[6]).strip()
                    section_name = str(row[7]).strip().upper()
                    year_name    = str(row[8]).strip()
                    father_name  = str(row[9]).strip() if row[9] else ''
                    mother_name  = str(row[10]).strip() if row[10] else ''
                    guardian_phone = str(row[11]).strip() if row[11] else ''
                    guardian_email = str(row[12]).strip() if row[12] else ''
                    adm_date     = row[13] if isinstance(row[13], date) else date.fromisoformat(str(row[13])) if row[13] else date.today()

                    division  = Division.objects.get(name=div_name)
                    grade     = Grade.objects.get(name=grade_name, division=division)
                    section   = Section.objects.get(name=section_name, grade=grade)
                    acad_year = AcademicYear.objects.get(name=year_name)

                    Student.objects.create(
                        full_name=full_name, arabic_name=arabic_name,
                        dob=dob, gender=gender, nationality=nationality,
                        division=division, grade=grade, section=section,
                        academic_year=acad_year,
                        father_name=father_name, mother_name=mother_name,
                        guardian_phone=guardian_phone, guardian_email=guardian_email,
                        admission_date=adm_date,
                        created_by=request.user,
                    )
                    created += 1
                except Exception as e:
                    errors.append(f"Row {i}: {e}")

            if errors:
                messages.warning(request, f"Imported {created} students. {len(errors)} rows had errors: " + " | ".join(errors[:5]))
            else:
                messages.success(request, f"Successfully imported {created} students.")

        except Exception as e:
            messages.error(request, f"Import failed: {e}")

        return redirect('students:list')

    columns = [
        {'name': 'full_name (EN)',                     'required': True,  'example': 'John Smith'},
        {'name': 'arabic_name (AR)',                   'required': True,  'example': 'جون سميث'},
        {'name': 'dob (YYYY-MM-DD)',                   'required': True,  'example': '2015-09-01'},
        {'name': 'gender (M/F)',                       'required': True,  'example': 'M'},
        {'name': 'nationality',                        'required': False, 'example': 'Saudi'},
        {'name': 'division (AMERICAN/BRITISH/FRENCH)', 'required': True,  'example': 'AMERICAN'},
        {'name': 'grade_name',                         'required': True,  'example': 'Grade 1'},
        {'name': 'section_name (A/B/C)',               'required': True,  'example': 'A'},
        {'name': 'academic_year',                      'required': True,  'example': '2024-25'},
        {'name': 'father_name',                        'required': False, 'example': 'Robert Smith'},
        {'name': 'mother_name',                        'required': False, 'example': 'Mary Smith'},
        {'name': 'guardian_phone',                     'required': False, 'example': '+966501234567'},
        {'name': 'guardian_email',                     'required': False, 'example': 'r.smith@email.com'},
        {'name': 'admission_date (YYYY-MM-DD)',         'required': False, 'example': '2025-09-01'},
    ]
    return render(request, 'students/student_import.html', {'columns': columns})


# ────────────────────────── EXCEL TEMPLATE DOWNLOAD ──────────────────────────

@login_required
@role_required(*_ADMIN)
def download_import_template(request):
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"
        headers = [
            'full_name (EN)', 'arabic_name (AR)', 'dob (YYYY-MM-DD)', 'gender (M/F)',
            'nationality', 'division (AMERICAN/BRITISH/FRENCH)', 'grade_name',
            'section_name (A/B/C)', 'academic_year (e.g. 2024-25)',
            'father_name', 'mother_name', 'guardian_phone', 'guardian_email',
            'admission_date (YYYY-MM-DD)',
        ]
        ws.append(headers)
        # Example row
        ws.append([
            'John Smith', 'جون سميث', '2015-09-01', 'M',
            'American', 'AMERICAN', 'Grade 1', 'A', '2024-25',
            'Robert Smith', 'Mary Smith', '+966501234567',
            'r.smith@email.com', '2025-09-01',
        ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
        return response
    except ImportError:
        messages.error(request, "openpyxl is not installed. Run: pip install openpyxl")
        return redirect('students:import')


# ────────────────────────── EXTERNAL CANDIDATE LIST ──────────────────────────

@login_required
@role_required(*_STAFF)
def external_list(request):
    """List all external exam candidates with search."""
    query = request.GET.get('q', '').strip()
    qs = ExternalCandidate.objects.select_related('grade_applying', 'board').annotate(
        payment_count=Count('payments'),
    )

    if query:
        qs = qs.filter(
            Q(full_name__icontains=query) |
            Q(candidate_id__icontains=query) |
            Q(arabic_name__icontains=query) |
            Q(phone__icontains=query) |
            Q(id_number__icontains=query)
        )

    return render(request, 'students/external_list.html', {
        'candidates': qs,
        'total':      qs.count(),
        'query':      query,
    })


# ────────────────────────── EXTERNAL CANDIDATE ADD ──────────────────────────

@login_required
@role_required(*_STAFF)
def external_add(request):
    """Register a new external candidate."""
    grades    = Grade.objects.select_related('division').order_by('division__name', 'order', 'name')
    divisions = Division.objects.filter(is_active=True).order_by('name')

    if request.method == 'POST':
        full_name    = request.POST.get('full_name', '').strip()
        arabic_name  = request.POST.get('arabic_name', '').strip()
        phone        = request.POST.get('phone', '').strip()
        nationality  = request.POST.get('nationality', '').strip()
        id_number    = request.POST.get('id_number', '').strip()
        grade_pk     = request.POST.get('grade_applying', '')
        division_pk  = request.POST.get('division', '')
        notes        = request.POST.get('notes', '').strip()
        is_saudi_val = request.POST.get('is_saudi', '')
        is_saudi     = True if is_saudi_val == 'saudi' else (False if is_saudi_val == 'non_saudi' else None)

        if not full_name:
            messages.error(request, "Full name is required.")
        else:
            grade_obj = None
            if grade_pk:
                try:
                    grade_obj = Grade.objects.get(pk=grade_pk)
                except Grade.DoesNotExist:
                    pass
            division_obj = None
            if division_pk:
                try:
                    division_obj = Division.objects.get(pk=division_pk)
                except Division.DoesNotExist:
                    pass

            candidate = ExternalCandidate.objects.create(
                full_name      = full_name,
                arabic_name    = arabic_name,
                phone          = phone,
                nationality    = nationality,
                id_number      = id_number,
                grade_applying = grade_obj,
                division       = division_obj,
                notes          = notes,
                is_saudi       = is_saudi,
                created_by     = request.user,
            )
            messages.success(request, f"Candidate {candidate.candidate_id} — {candidate.full_name} registered successfully.")
            return redirect('students:external_detail', pk=candidate.pk)

    return render(request, 'students/external_form.html', {
        'grades':    grades,
        'divisions': divisions,
        'candidate': None,
    })


# ────────────────────────── EXTERNAL CANDIDATE EDIT ──────────────────────────

@login_required
@role_required(*_STAFF)
def external_edit(request, pk):
    """Edit an existing external candidate."""
    candidate = get_object_or_404(ExternalCandidate, pk=pk)
    grades    = Grade.objects.select_related('division').order_by('division__name', 'order', 'name')
    divisions = Division.objects.filter(is_active=True).order_by('name')

    if request.method == 'POST':
        full_name    = request.POST.get('full_name', '').strip()
        arabic_name  = request.POST.get('arabic_name', '').strip()
        phone        = request.POST.get('phone', '').strip()
        nationality  = request.POST.get('nationality', '').strip()
        id_number    = request.POST.get('id_number', '').strip()
        grade_pk     = request.POST.get('grade_applying', '')
        division_pk  = request.POST.get('division', '')
        notes        = request.POST.get('notes', '').strip()
        is_saudi_val = request.POST.get('is_saudi', '')
        is_saudi     = True if is_saudi_val == 'saudi' else (False if is_saudi_val == 'non_saudi' else None)

        if not full_name:
            messages.error(request, "Full name is required.")
        else:
            grade_obj = None
            if grade_pk:
                try:
                    grade_obj = Grade.objects.get(pk=grade_pk)
                except Grade.DoesNotExist:
                    pass
            division_obj = None
            if division_pk:
                try:
                    division_obj = Division.objects.get(pk=division_pk)
                except Division.DoesNotExist:
                    pass

            candidate.full_name      = full_name
            candidate.arabic_name    = arabic_name
            candidate.phone          = phone
            candidate.nationality    = nationality
            candidate.id_number      = id_number
            candidate.grade_applying = grade_obj
            candidate.division       = division_obj
            candidate.notes          = notes
            candidate.is_saudi       = is_saudi
            candidate.save()
            messages.success(request, f"Candidate {candidate.full_name} updated.")
            return redirect('students:external_detail', pk=candidate.pk)

    return render(request, 'students/external_form.html', {
        'grades':    grades,
        'divisions': divisions,
        'candidate': candidate,
    })


# ────────────────────────── EXTERNAL CANDIDATE DETAIL ──────────────────────────

@login_required
@role_required(*_STAFF)
def external_detail(request, pk):
    """View details and payment history of an external candidate."""
    candidate = get_object_or_404(
        ExternalCandidate.objects.select_related('grade_applying', 'board'),
        pk=pk,
    )
    payments = ExternalCandidatePayment.objects.filter(
        candidate=candidate,
    ).order_by('-payment_date', '-id')

    total_paid = payments.aggregate(s=Sum('total'))['s'] or 0

    return render(request, 'students/external_detail.html', {
        'candidate':  candidate,
        'payments':   payments,
        'total_paid': total_paid,
    })

